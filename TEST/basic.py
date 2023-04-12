import csv
from openpyxl import Workbook

# 创建一个新的Workbook对象
wb = Workbook()

# 获取默认的活动工作表
ws = wb.active

# 在表格中添加内容
ws['A1'] = 1
ws['B1'] = 1
ws['A2'] = 1
ws['B2'] = 1

# 将表格保存为CSV文件
with open('example.csv', 'w', newline='') as f:
    c = csv.writer(f)
    for row in ws.iter_rows(values_only=True):
        c.writerow(row)
